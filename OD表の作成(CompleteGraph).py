#!/usr/bin/env python
# coding: utf-8

# In[3]:


import pandas as pd
import numpy as np
import openpyxl
import collections
import os

dft = pd.read_excel("prefecture_name.xlsx")
df1 = pd.read_csv("FF-data2016.csv",index_col=["サンプルID"], encoding="shift-jis")

#滞在日数の欠損値を０にする
df2 = df1.replace('-','0')
df3 = df2.assign(taizai = df2['滞在日数'].astype(np.int64))

#旅行目的の人とtripNoが1の人を抽出
df4 = df3.query("旅行目的コード == 1 & tripNo == 1")

#対象となる人のsampleIDを取得
b = [0 for i in range(len(df4))]
for i in range(len(df4)):
    b[i] = df4.iat[i,0]

#tripを求める
l = []
#都道府県のOD表を作成
a = [[0 for i in range(47)]for j in range(47)]

#tripに現れていない関係のある都道府県のトリップ量を追加
for i in b:
    df10 = df1.xs(i)
    df11 = df10.query("出発地コード < 50 & 目的地コード < 50 ")
    if len(df11) > 0:
        p = [[0 for i in range(47)]for j in range(47)]
        for j in range(len(df11)-1):
            p[df11.iat[j, 13]-1][df11.iat[j, 16]-1] += 1
            p[df11.iat[j, 16]-1][df11.iat[j, 13]-1] += 1
        for j in range(len(df11)-1):
            if df11.iat[j, 13] not in l:
                l.append(df11.iat[j, 13])
            if df11.iat[j, 16] not in l:
                l.append(df11.iat[j, 16])
        if len(l) >2:
            for j in range(len(l)-1):
                for k in range(j+1, len(l)):
                    if p[l[j]-1][l[k]-1] == 0 and p[l[k]-1][l[j]-1] == 0:
                        a[l[j]-1][l[k]-1] += sum([(df11.iat[0, p]) for p in range(29,37)])
    l.clear()

df5 = df3.query("旅行目的コード == 1 & 出発地コード < 50 & 目的地コード < 50 ")

#トリップチェーンの内、前後の都道府県のみに着目したときに現れる移動量
for i in range(len(df5)-1):
    a[df5.iloc[i,13]-1][df5.iloc[i,16]-1] += sum([(df5.iloc[i, p]) for p in range(29,37)])
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(47):
    cell = sheet.cell(row = 1, column = i+2)
    cell.value = dft.iloc[i,0]
for i in range(47):
    cell = sheet.cell(row = i+2, column = 1)
    cell.value = dft.iloc[i,0]
for i in range(47):
    for j in range(47):
        cell = sheet.cell(row = i+2, column = j+2)
        cell.value = str(round(a[i][j]))
wb.save('todouhuken_OD(Complete).xlsx')


# In[ ]:




